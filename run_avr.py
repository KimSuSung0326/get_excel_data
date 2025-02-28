import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# 파일 찾고 리스트에 저장하는 함수
def get_data_list(file_path):
    all_files = glob.glob(file_path)
    #print("CSV 파일들:", all_files)

    df_list = []
    split_file =[]
    for filename in all_files:
        df = pd.read_csv(filename)
        df_list.append(df)

        # 파일명만 추출 (경로 제외)
        base_name = os.path.basename(filename)
        # 추출한 파일명에서 호실만 분리
        split_file_name = "_".join(base_name.split("_")[:2])
        
        split_file.append(split_file_name)
        

        #print(f"파일 이름은 '{split_file}' 입니다.")
    return df_list,split_file


# 평균과 합을 계산하고 엑셀로 저장하는 함수
def save_avg_to_excel(df_list, output_file, room_num_list):
    result_data = []  # 결과를 저장할 리스트
    
    for df, room_name in zip(df_list, room_num_list):
        if not df.empty:
            column_name = df.columns[1]  # 두 번째 열 이름
            column_values = df[column_name]  # 해당 열 값

            # 0 값을 제외한 column_values 필터링
            filtered_values = column_values[column_values != 0]

            column_sum = filtered_values.sum()  # 합
            column_mean = round(filtered_values.mean())  
            
            conv_sum = int(column_sum)  # 정수 변환
            conv_mean = int(column_mean)

            # 데이터프레임에 추가할 데이터
            result_data.append({"파일명": room_name, "합계": conv_sum, "평균": conv_mean})

    # 결과 데이터프레임 생성
    result_df = pd.DataFrame(result_data)

    # 엑셀로 저장
    result_df.to_excel(output_file, index=False, sheet_name="심박 평균결과")

    #-----------------여기부터 엑셀 디자인 하는 코드-------------------
    wb = load_workbook(output_file)
    ws = wb['결과']

    # 셀 스타일을 적용할 경계선 스타일 설정
    thick_border = Border(
        top=Side(style='thick', color='000000'),
        #bottom=Side(style='thick', color='000000'),
        #left=Side(style='thick', color='000000'),
        #right=Side(style='thick', color='000000')
    )
      
    
    # 처음 호실 index 값 추출
    for i in range(1):
        prev_room_prefix = room_num_list[0].split("_")[0]
        #print(f"처음 인덱스 값은 '{prev_room_prefix}' 입니다")

    for row_idx, room_index in enumerate(room_num_list, start=2):  # start=2는 Excel에서 첫 번째 행을 1로부터 시작하도록
        room_prefix = room_index.split("_")[0]  # '311', '320' 부분 추출

        # 호수가 바뀌면 경계선 추가
        if prev_room_prefix != room_prefix:
            for cell in ws[row_idx]:  # 해당 행에 있는 모든 셀에 경계선 적용
                cell.border = thick_border

        # 이전 호수 값 갱신
        prev_room_prefix = room_prefix

# 변경된 엑셀 파일 저장
    wb.save(output_file)
       


    #print(f"결과가 '{output_file}' 파일로 저장되었습니다.")


# ----------------------------------------------
# main 함수 호출
if __name__ == "__main__":
    file_path = input("CSV 파일 경로를 입력하세요: ")
    file_path = file_path + r"\*.csv"  # 경로 수정

    data_list, room_num_list = get_data_list(file_path)

    # 평균과 합을 저장할 엑셀 파일 이름 지정
    output_excel_file = "심박_평균_합계.xlsx"
    save_avg_to_excel(data_list, output_excel_file,room_num_list)
